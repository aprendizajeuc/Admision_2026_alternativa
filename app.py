import streamlit as st
import pandas as pd
import io
import json
from datetime import datetime
import os
from openai import OpenAI
import PyPDF2
import docx2txt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading

st.set_page_config(
    page_title="Sistema de Analisis de Admision - SDT",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed"
)

st.markdown("""
<style>
    :root {
        --primary-blue: #1E3A8A;
        --secondary-blue: #3B82F6;
        --success-green: #059669;
        --warning-orange: #F59E0B;
        --error-red: #DC2626;
        --bg-light: #F8FAFC;
        --text-dark: #0F172A;
        --text-medium: #334155;
        --text-light: #64748B;
        --border-color: #E2E8F0;
    }
    .main, .stApp, [data-testid="stAppViewContainer"], [data-testid="stMain"],
    [data-testid="stMainBlockContainer"], [data-testid="stVerticalBlock"],
    [data-testid="stAppViewBlockContainer"], section[data-testid="stSidebar"],
    .block-container {
        background-color: #F1F5F9 !important; color: #0F172A !important;
    }
    .main { background: linear-gradient(180deg, #EFF6FF 0%, #F1F5F9 100%) !important; padding: 2rem 1rem; }
    div[data-testid="stFileUploader"] { background: #FFFFFF !important; border-radius: 12px; padding: 2rem; border: 2px dashed #3B82F6 !important; box-shadow: 0 4px 6px rgba(0,0,0,0.05); }
    div[data-testid="stFileUploader"]:hover { border-color: #1E3A8A !important; }
    div[data-testid="stFileUploader"] label, div[data-testid="stFileUploader"] small,
    div[data-testid="stFileUploader"] span, div[data-testid="stFileUploader"] p,
    div[data-testid="stFileUploader"] div { color: #0F172A !important; font-weight: 500 !important; }
    div[data-testid="stFileUploader"] section, div[data-testid="stFileUploader"] section > div,
    [data-testid="stFileUploaderDropzone"] { background-color: #FFFFFF !important; background: #FFFFFF !important; color: #0F172A !important; border-color: #3B82F6 !important; }
    div[data-testid="stFileUploader"] button, [data-testid="stFileUploaderDropzone"] button { background-color: #3B82F6 !important; color: #FFFFFF !important; border: none !important; font-weight: 600 !important; }
    div[data-testid="stFileUploader"] button:hover { background-color: #1E3A8A !important; }
    [data-testid="stFileUploaderDropzoneInstructions"] div, [data-testid="stFileUploaderDropzoneInstructions"] span,
    [data-testid="stFileUploaderDropzoneInstructions"] small { color: #334155 !important; }
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"],
    div[data-testid="stFileUploader"] [data-testid="stFileUploaderFile"] * { background-color: #F8FAFC !important; color: #0F172A !important; }
    .app-header { background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%); padding: 2rem; border-radius: 16px; margin-bottom: 2rem; box-shadow: 0 10px 30px rgba(30,58,138,0.2); text-align: center; }
    .app-title { font-size: 2.5rem; font-weight: 800; color: #FFFFFF !important; margin: 0; }
    .app-subtitle { font-size: 1.1rem; color: #F1F5F9 !important; margin-top: 0.5rem; }
    .stButton > button { background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%); color: #FFFFFF !important; border: none; padding: 0.75rem 2rem; font-size: 1rem; font-weight: 600; border-radius: 8px; box-shadow: 0 4px 12px rgba(30,58,138,0.3); }
    .stButton > button:hover { transform: translateY(-2px); box-shadow: 0 6px 20px rgba(30,58,138,0.4); }
    div[data-testid="metric-container"] { background: #FFFFFF !important; padding: 1.5rem; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border-left: 4px solid #3B82F6; }
    div[data-testid="metric-container"] label { color: #334155 !important; font-size: 0.875rem; font-weight: 600; }
    div[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #1E3A8A !important; font-size: 2rem; font-weight: 700; }
    .score-circle { display: inline-flex; width: 100px; height: 100px; border-radius: 50%; background: linear-gradient(135deg, #1E3A8A 0%, #3B82F6 100%); color: #FFFFFF !important; align-items: center; justify-content: center; font-size: 2.2rem; font-weight: 800; box-shadow: 0 8px 20px rgba(30,58,138,0.3); position: relative; }
    .score-circle::after { content: '/20'; position: absolute; bottom: 10px; right: 10px; font-size: 0.7rem; opacity: 0.9; color: #FFFFFF !important; }
    .stProgress > div > div > div > div { background: linear-gradient(90deg, #1E3A8A 0%, #3B82F6 100%); }
    div[data-testid="stExpander"] { background-color: #FFFFFF !important; border: 1px solid #E2E8F0 !important; border-radius: 8px !important; margin-bottom: 0.5rem; overflow: hidden; }
    div[data-testid="stExpander"] details > summary { background-color: #FFFFFF !important; color: #0F172A !important; font-weight: 600; padding: 0.75rem 1rem; border: none !important; }
    div[data-testid="stExpander"] details > summary:hover { background-color: #F8FAFC !important; }
    div[data-testid="stExpander"] details[open] > summary { background-color: #FFFFFF !important; color: #0F172A !important; border-bottom: 2px solid #3B82F6 !important; }
    div[data-testid="stExpander"] details > summary *, div[data-testid="stExpander"] summary [data-testid="stMarkdownContainer"] * { color: #0F172A !important; }
    div[data-testid="stExpander"] details > div, div[data-testid="stExpander"] details[open] > div,
    div[data-testid="stExpander"] details[open] > div > div, [data-testid="stExpanderDetails"] { background-color: #FFFFFF !important; background: #FFFFFF !important; color: #0F172A !important; }
    div[data-testid="stExpander"] details[open] p, div[data-testid="stExpander"] details[open] div,
    div[data-testid="stExpander"] details[open] span, div[data-testid="stExpander"] details[open] strong,
    [data-testid="stExpanderDetails"] *, [data-testid="stExpanderDetails"] p,
    [data-testid="stExpanderDetails"] div, [data-testid="stExpanderDetails"] span { color: #0F172A !important; background-color: transparent !important; }
    div[data-testid="stExpander"] [data-testid="stAlert"], [data-testid="stExpanderDetails"] [data-testid="stAlert"] { background-color: #DBEAFE !important; }
    div[data-testid="stExpander"] [data-testid="stAlert"] *, [data-testid="stExpanderDetails"] [data-testid="stAlert"] * { color: #1E3A8A !important; }
    .stDownloadButton > button { background: linear-gradient(135deg, #059669 0%, #10B981 100%); color: #FFFFFF !important; border: none; padding: 0.75rem 2rem; font-size: 1rem; font-weight: 600; border-radius: 8px; box-shadow: 0 4px 12px rgba(5,150,105,0.3); }
    .stDownloadButton > button:hover { transform: translateY(-2px); }
    .stAlert { border-radius: 8px; border-left: 4px solid; }
    [data-testid="stAlert"][data-baseweb*="positive"], .stSuccess { background-color: #D1FAE5 !important; }
    [data-testid="stAlert"][data-baseweb*="positive"] *, .stSuccess * { color: #065F46 !important; }
    [data-testid="stAlert"][data-baseweb*="info"], .stInfo { background-color: #DBEAFE !important; }
    [data-testid="stAlert"][data-baseweb*="info"] *, .stInfo * { color: #1E3A8A !important; }
    [data-testid="stAlert"][data-baseweb*="warning"], .stWarning { background-color: #FEF3C7 !important; }
    .stWarning * { color: #92400E !important; }
    [data-testid="stAlert"][data-baseweb*="negative"], .stError { background-color: #FEE2E2 !important; }
    .stError * { color: #991B1B !important; }
    .status-badge { display: inline-block; padding: 0.4rem 1rem; border-radius: 20px; font-weight: 600; font-size: 0.875rem; }
    .badge-success { background: #D1FAE5; color: #065F46 !important; }
    .badge-warning { background: #FEF3C7; color: #92400E !important; }
    .badge-error { background: #FEE2E2; color: #991B1B !important; }
    .badge-info { background: #DBEAFE; color: #1E40AF !important; }
    h1 { color: #0F172A !important; font-weight: 800 !important; }
    h2 { color: #0F172A !important; font-weight: 700 !important; margin-top: 2rem !important; }
    h3 { color: #1E3A8A !important; font-weight: 600 !important; }
    h4 { color: #0F172A !important; font-weight: 600 !important; }
    p, li, span, div, label { color: #0F172A !important; }
    small, .stCaption { color: #334155 !important; }
    .stMarkdown, .stMarkdown p, .stMarkdown div, .stMarkdown span { color: #0F172A !important; }
    strong, b { color: #1E3A8A !important; font-weight: 700 !important; }
    hr { border: none; height: 2px; background: linear-gradient(90deg, transparent, #CBD5E1, transparent); margin: 2rem 0; }
    .dataframe th { background-color: #1E3A8A !important; color: #FFFFFF !important; font-weight: 600 !important; }
    .dataframe td { color: #0F172A !important; background-color: #FFFFFF !important; }
    .block-container { padding-top: 2rem; padding-bottom: 2rem; max-width: 1400px; }
    .info-box { background: #DBEAFE; border-left: 4px solid #3B82F6; padding: 1rem; border-radius: 8px; margin: 1rem 0; }
    .info-box *, .info-box strong, .info-box span, .info-box p { color: #1E3A8A !important; }
    input, textarea, select { color: #0F172A !important; background-color: #FFFFFF !important; }
    @media (max-width: 768px) { .app-title { font-size: 1.8rem; } .score-circle { width: 80px; height: 80px; font-size: 1.8rem; } }
</style>
""", unsafe_allow_html=True)

@st.cache_resource
def get_openai_client():
    api_key = None
    try:
        api_key = st.secrets["OPENAI_API_KEY"]
    except:
        api_key = os.getenv('OPENAI_API_KEY')
    if not api_key:
        st.error("No se encontro la clave API de OpenAI")
        st.info("**Para uso local:** Configura OPENAI_API_KEY en .streamlit/secrets.toml")
        st.info("**Para Streamlit Cloud:** Configura OPENAI_API_KEY en Settings -> Secrets")
        st.stop()
    return OpenAI(api_key=api_key)

client = get_openai_client()

def find_column(df, variants):
    """Busca columna por nombre exacto (case-insensitive)."""
    df_cols_lower = {col.strip().lower(): col for col in df.columns}
    for variant in variants:
        key = variant.strip().lower()
        if key in df_cols_lower:
            return df_cols_lower[key]
    return None

def find_columns_by_prefix(df, prefixes):
    """Busca columnas que EMPIECEN con alguno de los prefijos dados. Retorna lista ordenada."""
    found = []
    for col in df.columns:
        col_lower = col.strip().lower()
        for prefix in prefixes:
            if col_lower.startswith(prefix.strip().lower()):
                found.append(col)
                break
    return found

def build_column_map(df):
    mappings = {
        'nombre': [
            'Nombre', 'Nombres', 'NOMBRE', 'NOMBRES',
            'nombre', 'nombres', 'Name', 'Primer Nombre',
            'nombre completo', 'Nombre Completo'
        ],
        'apellidos': [
            'Apellido(s)', 'Apellidos', 'Apellido',
            'APELLIDO(S)', 'APELLIDOS', 'APELLIDO',
            'apellido(s)', 'apellidos', 'apellido',
            'Last Name', 'Surname', 'Apellido Paterno'
        ],
        'correo': [
            'Direccion de correo', 'Dirección de correo',
            'DIRECCION DE CORREO', 'DIRECCIÓN DE CORREO',
            'direccion de correo', 'dirección de correo',
            'Correo electronico', 'Correo Electronico',
            'Correo electrónico', 'Correo Electrónico',
            'correo electronico', 'correo electrónico',
            'Correo', 'correo', 'CORREO',
            'Email', 'email', 'EMAIL', 'E-mail', 'Mail', 'mail'
        ],
        'edad': ['Edad', 'edad', 'EDAD', 'Age'],
        'programa': [
            'Programa', 'programa', 'PROGRAMA',
            'Carrera', 'carrera', 'CARRERA',
            'Programa Academico', 'Programa Académico',
            'Especialidad', 'especialidad', 'Facultad'
        ],
        'respuesta_1': [
            'Respuesta 1', 'respuesta 1', 'RESPUESTA 1',
            'Respuesta1', 'R1', 'r1', 'Pregunta 1', 'P1'
        ],
        'respuesta_2': [
            'Respuesta 2', 'respuesta 2', 'RESPUESTA 2',
            'Respuesta2', 'R2', 'r2', 'Pregunta 2', 'P2'
        ],
        'respuesta_3': [
            'Respuesta 3', 'respuesta 3', 'RESPUESTA 3',
            'Respuesta3', 'R3', 'r3', 'Pregunta 3', 'P3'
        ],
    }
    col_map = {}
    for logical_name, variants in mappings.items():
        col_map[logical_name] = find_column(df, variants)
    
    # Fallback: si no se encontraron respuestas por nombre exacto,
    # buscar columnas que empiecen con "Comentario" (formato Moodle/LMS)
    respuestas_faltantes = (
        col_map['respuesta_1'] is None or
        col_map['respuesta_2'] is None or
        col_map['respuesta_3'] is None
    )
    
    if respuestas_faltantes:
        comentario_cols = find_columns_by_prefix(df, ['Comentario -', 'Comentario-', 'Comentario'])
        
        if len(comentario_cols) >= 3:
            if col_map['respuesta_1'] is None:
                col_map['respuesta_1'] = comentario_cols[0]
            if col_map['respuesta_2'] is None:
                col_map['respuesta_2'] = comentario_cols[1]
            if col_map['respuesta_3'] is None:
                col_map['respuesta_3'] = comentario_cols[2]
        elif len(comentario_cols) > 0:
            # Asignar las que haya en orden
            idx = 0
            for key in ['respuesta_1', 'respuesta_2', 'respuesta_3']:
                if col_map[key] is None and idx < len(comentario_cols):
                    col_map[key] = comentario_cols[idx]
                    idx += 1
    
    return col_map

def safe_get(row, col_name, default='N/A'):
    if col_name is None:
        return default
    val = row.get(col_name, default)
    if pd.isna(val) or str(val).strip() == '':
        return default
    return str(val).strip()

def extract_text_from_pdf(file):
    try:
        pdf_reader = PyPDF2.PdfReader(file)
        text = ""
        for page in pdf_reader.pages:
            text += page.extract_text()
        return text
    except Exception as e:
        st.error(f"Error al leer PDF: {str(e)}")
        return None

def extract_text_from_docx(file):
    try:
        return docx2txt.process(file)
    except Exception as e:
        st.error(f"Error al leer DOCX: {str(e)}")
        return None

def extract_text_from_txt(file):
    try:
        return file.read().decode('utf-8')
    except Exception as e:
        st.error(f"Error al leer TXT: {str(e)}")
        return None

def read_excel_file(file):
    try:
        ext = file.name.split('.')[-1].lower()
        return pd.read_csv(file) if ext == 'csv' else pd.read_excel(file)
    except Exception as e:
        st.error(f"Error al leer archivo: {str(e)}")
        return None


# =====================================================================
# PROMPT v4 - Preguntas nuevas + rubrica adaptada + sub-dimensiones
# =====================================================================

SYSTEM_PROMPT = """ROL: Experto en Psicologia Educativa especializado en Teoria de la Autodeterminacion (SDT) de Ryan y Deci.

CONTEXTO: Universidad Continental Peru, modalidad a distancia. Poblacion diversa, 18+ anios. Objetivo: diagnosticar motivacion, NO evaluar ortografia ni redaccion.

PREGUNTAS DEL FORMULARIO:
P1 - PROCESO DE ELECCION: Cuentanos como fue tu proceso para elegir la carrera universitaria a la que postulas. Que hiciste?, que pensabas? y como te sentias?
P2 - EXPERIENCIA PERSONAL: Cuentanos una experiencia personal de cualquier etapa de tu vida, en la que hayas realizado algo que se relacione a la carrera que postulas. Que hiciste?, que pensabas? y como te sentiste?
P3 - PROYECCION DE VIDA A 10 ANIOS: Imagina que ya han pasado 10 anios desde tu graduacion. Cuentanos: como serian tus dias?, en que actividades estarias involucrado(a)? y por que?

ESCALA SDT (1-6):
6=Intrinseca | 5=Integrada | 4=Identificada | 3=Introyectada | 2=Externa | 1=Amotivacion

=====================================================================
SUB-DIMENSIONES DE ANALISIS
=====================================================================

Cada respuesta contiene hasta 3 capas de informacion. Extrae indicadores de CADA capa antes de asignar un nivel:

CAPA 1 - ACCIONES (que hizo): Comportamientos concretos narrados. Indaga, explora, investiga (autonomo) vs. alguien le dijo, le toco, no hizo nada (pasivo/externo).

CAPA 2 - PENSAMIENTOS (que pensaba): Razonamientos y creencias. Reflexiona sobre valor personal (identificado) vs. piensa en dinero/estatus (externo) vs. piensa que debe cumplir (introyectado).

CAPA 3 - EMOCIONES (como se sentia): Estados emocionales. Disfrute/curiosidad (intrinseca) vs. verguenza/culpa/orgullo (introyectada) vs. indiferencia (amotivacion).

REGLA DE INTEGRACION: El nivel final de cada respuesta se determina por el indicador MAS BAJO encontrado en CUALQUIERA de las 3 capas.

=====================================================================
RUBRICA ADAPTADA A LAS NUEVAS PREGUNTAS
=====================================================================

--- P1: PROCESO DE ELECCION DE CARRERA ---

NIVEL 6 - INTRINSECA: El proceso de eleccion estuvo guiado por curiosidad, fascinacion o disfrute genuino por los contenidos o actividades de la carrera. Busco informacion por interes propio, pensaba en lo mucho que le atraian los temas, se sentia entusiasmado/a al explorar.
NIVEL 5 - INTEGRADA: La eleccion surge de una reflexion profunda sobre quien es y quien quiere ser. El proceso revela coherencia con valores centrales e identidad. Sentia que la carrera encaja con su esencia.
NIVEL 4 - IDENTIFICADA: Eligio porque reconoce el valor e importancia de la carrera para sus metas personales o para contribuir. Investigo, evaluo opciones y decidio con conviccion por utilidad significativa.
NIVEL 3 - INTROYECTADA: El proceso estuvo marcado por presion emocional interna: verguenza, culpa, miedo a decepcionar, necesidad de demostrar. Elegia para complacer, para no sentirse mal, para tener imagen ante otros.
NIVEL 2 - EXTERNA: Eligio por factores tangibles externos: salario, empleabilidad, estabilidad, prestigio, demanda laboral. El proceso fue guiado por lo que la carrera ofrece materialmente.
NIVEL 1 - AMOTIVACION: No hubo proceso real de eleccion. Elegia por inercia, sin razon clara, porque alguien decidio por el/ella, o no sabe por que postula.

--- P2: EXPERIENCIA PERSONAL RELACIONADA ---

NIVEL 6 - INTRINSECA: La experiencia fue disfrutada por si misma. Narra el proceso con entusiasmo, curiosidad, sensacion de flujo. Pensaba en lo interesante de la actividad. Se sentia absorto/a, motivado/a, curioso/a.
NIVEL 5 - INTEGRADA: La experiencia confirmo o fortalecio su identidad. Descubrio que la actividad es parte de quien es. Pensaba en coherencia con sus valores. Se sentia pleno/a, autentico/a, alineado/a.
NIVEL 4 - IDENTIFICADA: Valoro la experiencia porque le enseno habilidades importantes o le mostro que puede contribuir. Pensaba en lo util del aprendizaje. Se sentia capaz, productivo/a.
NIVEL 3 - INTROYECTADA: La experiencia estuvo marcada por autoexigencia, necesidad de demostrar, miedo a fallar. Pensaba en no quedar mal o en probar que es capaz. Se sentia presionado/a, aliviado/a al terminar, orgulloso/a de no fallar.
NIVEL 2 - EXTERNA: La experiencia fue valorada por sus resultados tangibles: nota, reconocimiento, premio, beneficio material. Pensaba en obtener el resultado. Se sentia satisfecho/a por la recompensa obtenida.
NIVEL 1 - AMOTIVACION: No tiene experiencia relevante, o la narra con indiferencia total. No penso nada especial. No sintio nada.

--- P3: PROYECCION DE VIDA A 10 ANIOS ---

NIVEL 6 - INTRINSECA: Se imagina sus dias realizando actividades de la carrera por DISFRUTE. Sus actividades estan centradas en el placer de hacer, explorar, aprender. El "por que" es porque le fascina, lo disfruta.
NIVEL 5 - INTEGRADA: Se imagina viviendo de manera COHERENTE CON SU IDENTIDAD. Sus dias reflejan quien quiere ser. Las actividades son expresion de sus valores. El "por que" es porque asi es el/ella.
NIVEL 4 - IDENTIFICADA: Se imagina ejerciendo con PROPOSITO y contribuyendo. Sus dias incluyen actividades profesionales significativas. El "por que" es porque valora lo que hace y quiere aportar.
NIVEL 3 - INTROYECTADA: Se imagina cumpliendo para NO FALLAR o para que otros lo validen. Sus actividades buscan demostrar exito. El "por que" es para sentirse a la altura, no decepcionar.
NIVEL 2 - EXTERNA: Se imagina con ESTABILIDAD MATERIAL. Sus dias giran en torno a empleo estable, buen sueldo, beneficios. El "por que" es por dinero, estatus, seguridad economica.
NIVEL 1 - AMOTIVACION: No se imagina ejerciendo, o su proyeccion es vaga/desconectada de la carrera. No sabe que haria. El "por que" no existe o es incoherente.

=====================================================================
REGLAS DE DISCRIMINACION OBLIGATORIAS (D1-D4)
=====================================================================

--- REGLA D1: Discriminar NIVEL 2 vs NIVEL 3 ---
Pregunta clave: "Que busca la persona: un RESULTADO TANGIBLE o resolver/gestionar una EMOCION?"
TANGIBLE (dinero, empleo, notas, estatus, reconocimiento, estabilidad) = NIVEL 2
EMOCION (verguenza, culpa, orgullo como motor, miedo a decepcionar, complacer) = NIVEL 3

Principios:
- Lo que define el nivel NO es quien aparece en la frase, sino QUE TIPO DE COSA motiva.
- Estatus social, respeto ajeno y dinero como motivadores = tangibles = NIVEL 2
- Verguenza ante otros, evitar decepcionar, buscar que se sientan orgullosos = emociones = NIVEL 3
- Buscar buenas notas/reconocimiento y luego sentirse satisfecho = satisfaccion es consecuencia de lo tangible = NIVEL 2
- Elegir algo para que otra persona sea feliz o no sufra = complacer emocionalmente = NIVEL 3
- Elegir algo por imagen/presencia motivado por verguenza = emocion interna = NIVEL 3

--- REGLA D2: Discriminar NIVEL 3 vs NIVEL 4 ---
Pregunta clave: "La persona VALORA el aprendizaje/actividad, o actua por OBLIGACION EMOCIONAL?"
VALORA (reconoce importancia, ve utilidad, quiere contribuir, aprecia aprendizaje) = NIVEL 4
OBLIGACION (sentir que cumple, demostrar capacidad, no fallar, no decepcionar) = NIVEL 3

Principios:
- En nivel 4 la persona ELIGE porque ve valor.
- En nivel 3 se SIENTE OBLIGADA emocionalmente.
- "Quiero contribuir a mejorar mi entorno" = NIVEL 4
- "Necesito sentir que cumplo con lo esperado" = NIVEL 3

--- REGLA D3: Discriminar NIVEL 4 vs NIVEL 5 ---
Pregunta clave: "Habla de METAS/UTILIDAD o de IDENTIDAD/PROYECTO DE VIDA?"
Metas, utilidad, desarrollo profesional, contribucion = NIVEL 4
Identidad, proyecto vital, coherencia con valores centrales, "quien quiero ser" = NIVEL 5

--- REGLA D4: Discriminar NIVEL 5 vs NIVEL 6 ---
Pregunta clave: "El foco esta en QUIEN SOY (identidad) o en lo que DISFRUTO HACER (proceso)?"
Coherencia con valores, identidad, proyecto de vida = NIVEL 5
Disfrute del proceso, curiosidad, placer en la actividad misma = NIVEL 6

=====================================================================
REGLAS ESPECIALES
=====================================================================

SATISFACCION Y RECOMPENSA EXTERNA (nivel 3 vs 2):
No clasificar como introyectada cuando el texto exprese unicamente busqueda de recompensa externa (nota, reconocimiento, premio), aun cuando mencione satisfaccion posterior. La satisfaccion derivada de una recompensa externa NO constituye presion interna.

RECOMPENSA EXTERNA CON EMOCION POSITIVA (nivel 2):
Cuando haya recompensa externa explicita + emocion positiva posterior, se asigna regulacion externa (2), SALVO que exista mencion explicita de culpa, verguenza, orgullo como motor, complacer a otros o autoexigencia.

=====================================================================
PROCESO DE EVALUACION (SEGUIR EN ORDEN ESTRICTO)
=====================================================================

Para CADA respuesta (P1, P2, P3):
1. Leer la respuesta completa sin prejuicios
2. Extraer indicadores de las 3 CAPAS: acciones, pensamientos, emociones
3. Clasificar cada indicador segun la rubrica ESPECIFICA de esa pregunta
4. Si hay indicadores de MULTIPLES niveles -> asignar el nivel MAS BAJO
5. VERIFICAR aplicando la regla D1, D2, D3 o D4 segun los niveles en juego
6. Si hay recompensa externa + satisfaccion posterior -> verificar REGLA ESPECIAL
7. Escribir justificacion citando evidencia de las 3 capas + regla aplicada

PERFIL FINAL:
- Base: Perfil = min(P1, P2, P3)
- Excepcion 2-de-3: Si dos puntajes coinciden y el tercero esta EXACTAMENTE 1 nivel abajo -> perfil = nivel coincidente
  Validos: (4,4,3)->4, (5,5,4)->5, (6,6,5)->6
  NO validos: (6,5,4)->4, (5,5,3)->3
- Seguridad: Si alguno = 1 -> perfil maximo = 2

CALCULOS:
- calificacion_real = P1+P2+P3 (max 18)
- calificacion_sobre_20 = (real/18)*20 con 2 decimales

=====================================================================
JSON DE RESPUESTA (responder SOLO con este JSON, sin markdown)
=====================================================================

{
  "informacion_extraida": {
    "nombre": "...", "apellidos": "...", "edad": "...", "programa": "...", "correo": "..."
  },
  "evaluacion_motivacional": {
    "eleccion_carrera": {
      "puntaje": 1-6,
      "tipo_motivacion": "nombre del nivel",
      "justificacion": "Acciones: [...]. Pensamientos: [...]. Emociones: [...]. Regla aplicada: [D1/D2/D3/D4 + explicacion]"
    },
    "experiencia_relacionada": {
      "puntaje": 1-6,
      "tipo_motivacion": "nombre del nivel",
      "justificacion": "Acciones: [...]. Pensamientos: [...]. Emociones: [...]. Regla aplicada: [D1/D2/D3/D4 + explicacion]"
    },
    "proyeccion_vida": {
      "puntaje": 1-6,
      "tipo_motivacion": "nombre del nivel",
      "justificacion": "Actividades imaginadas: [...]. Razon (por que): [...]. Regla aplicada: [D1/D2/D3/D4 + explicacion]"
    }
  },
  "necesidades_psicologicas": {
    "autonomia": "Alta/Media/Baja - Analisis breve",
    "competencia": "Alta/Media/Baja - Analisis breve",
    "relacion": "Alta/Media/Baja - Analisis breve"
  },
  "calificacion_real": 0,
  "calificacion_sobre_20": 0.00,
  "perfil_motivacional_final": "...",
  "regla_aplicada": "min(X,Y,Z)=N o excepcion aplicada",
  "recomendaciones": "...",
  "nivel_motivacional_general": "Predominantemente ..."
}"""

USER_PROMPT_TEMPLATE = """PREGUNTAS DEL FORMULARIO:
1. Cuentanos como fue tu proceso para elegir la carrera universitaria a la que postulas. Que hiciste?, que pensabas? y como te sentias?
2. Cuentanos una experiencia personal de cualquier etapa de tu vida, en la que hayas realizado algo que se relacione a la carrera que postulas. Que hiciste?, que pensabas? y como te sentiste?
3. Imagina que ya han pasado 10 anios desde tu graduacion. Cuentanos: como serian tus dias?, en que actividades estarias involucrado(a)? y por que?

FORMULARIO:
{text_content}

INSTRUCCION FINAL:
1) Lee cada respuesta completa
2) Extrae indicadores de las 3 CAPAS: acciones, pensamientos, emociones
3) Si hay multiples niveles -> asigna el nivel inferior
4) APLICA la regla de discriminacion D1/D2/D3/D4 segun corresponda
5) Si hay recompensa externa + satisfaccion posterior -> verifica REGLA ESPECIAL nivel 2
6) Justifica con evidencia de las 3 capas + regla aplicada
7) Calcula perfil final
8) Responde SOLO con JSON valido (sin markdown, sin backticks)"""


def analyze_admission_form(text_content, retry_count=0):
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": USER_PROMPT_TEMPLATE.format(text_content=text_content)}
            ],
            temperature=0.1,
            max_tokens=3500
        )
        
        content = response.choices[0].message.content.strip()
        
        if content.startswith('```'):
            lines = content.split('\n')
            content = '\n'.join(lines[1:-1]) if len(lines) > 2 else content
            content = content.replace('```json', '').replace('```', '').strip()
        
        result = json.loads(content)
        result['tokens_used'] = response.usage.total_tokens
        result['timestamp'] = datetime.now().isoformat()
        result['success'] = True
        return result
        
    except json.JSONDecodeError as e:
        if retry_count < 2:
            import time
            time.sleep(1)
            return analyze_admission_form(text_content, retry_count + 1)
        return {"success": False, "error": f"Error JSON tras {retry_count + 1} intentos", "detail": str(e)}
    except Exception as e:
        return {"success": False, "error": str(e)}


def process_excel_records(df, progress_bar, status_text):
    total = len(df)
    col_map = build_column_map(df)
    
    missing = {k for k, v in col_map.items() if v is None}
    if missing:
        st.warning(f"Columnas no encontradas: **{', '.join(missing)}**. Se usara 'N/A'.\n\n**Columnas en archivo:** {', '.join(df.columns.tolist())}")
    
    # Preparar todos los registros antes de procesar
    tasks = []
    skipped = []
    
    for idx, row in df.iterrows():
        nombre = safe_get(row, col_map['nombre'])
        apellidos = safe_get(row, col_map['apellidos'])
        correo = safe_get(row, col_map['correo'])
        edad = safe_get(row, col_map['edad'])
        programa = safe_get(row, col_map['programa'])
        resp1 = safe_get(row, col_map['respuesta_1'], 'Sin respuesta')
        resp2 = safe_get(row, col_map['respuesta_2'], 'Sin respuesta')
        resp3 = safe_get(row, col_map['respuesta_3'], 'Sin respuesta')
        
        form_text = f"""Nombre: {nombre}
Apellidos: {apellidos}
Correo: {correo}
Edad: {edad}
Programa: {programa}

Pregunta 1 - Proceso de eleccion de carrera:
{resp1}

Pregunta 2 - Experiencia personal relacionada:
{resp2}

Pregunta 3 - Proyeccion de vida a 10 anios:
{resp3}"""
        
        missing_responses = []
        if resp1 == 'Sin respuesta': missing_responses.append('Respuesta 1')
        if resp2 == 'Sin respuesta': missing_responses.append('Respuesta 2')
        if resp3 == 'Sin respuesta': missing_responses.append('Respuesta 3')
        
        base_info = {
            'registro_numero': idx + 1,
            'nombre': nombre,
            'apellidos': apellidos,
            'correo': correo,
        }
        
        if missing_responses:
            skipped.append({**base_info, 'success': False, 'error': f"Campos faltantes: {', '.join(missing_responses)}"})
        else:
            tasks.append((base_info, form_text))
    
    # Contador thread-safe para progreso
    completed_count = [0]
    lock = threading.Lock()
    total_to_process = len(tasks) + len(skipped)
    
    def process_single(task):
        base_info, form_text = task
        analysis = analyze_admission_form(form_text)
        result = {**base_info, 'success': analysis.get('success', False)}
        if analysis.get('success'):
            result['analysis'] = analysis
        else:
            result['error'] = analysis.get('error', 'Error desconocido')
        
        with lock:
            completed_count[0] += 1
        
        return result
    
    # Procesar en paralelo con 5 hilos
    parallel_results = []
    max_workers = min(5, len(tasks)) if tasks else 1
    
    if tasks:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(process_single, task): task[0]['registro_numero'] for task in tasks}
            
            for future in as_completed(futures):
                result = future.result()
                parallel_results.append(result)
                
                done = completed_count[0] + len(skipped)
                progress_bar.progress(done / total_to_process)
                status_text.markdown(f"**Procesando:** {done} de {total_to_process} registros ({max_workers} en paralelo)")
    
    # Combinar resultados: skipped + procesados, ordenar por numero de registro
    all_results = skipped + parallel_results
    all_results.sort(key=lambda x: x['registro_numero'])
    
    return all_results


def generate_excel_report(results):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados Analisis SDT"
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['N', 'Nombre', 'Apellidos', 'Correo', 'Calif. Real', 'Calif. /20',
        'R1 Punt.', 'R1 Justificacion', 'R1 Tipo', 'R2 Punt.', 'R2 Justificacion', 'R2 Tipo',
        'R3 Punt.', 'R3 Justificacion', 'R3 Tipo', 'Nivel General', 'Autonomia', 'Competencia', 'Relacion', 'Recomendaciones']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    widths = [5, 15, 15, 25, 10, 10, 8, 50, 15, 8, 50, 15, 8, 50, 15, 20, 30, 30, 30, 50]
    for col_num, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    for result in results:
        a = result.get('analysis', {})
        ws.append([
            result.get('registro_numero', ''), result.get('nombre', ''), result.get('apellidos', ''), result.get('correo', ''),
            a.get('calificacion_real', ''), a.get('calificacion_sobre_20', ''),
            a.get('evaluacion_motivacional', {}).get('eleccion_carrera', {}).get('puntaje', ''),
            a.get('evaluacion_motivacional', {}).get('eleccion_carrera', {}).get('justificacion', ''),
            a.get('evaluacion_motivacional', {}).get('eleccion_carrera', {}).get('tipo_motivacion', ''),
            a.get('evaluacion_motivacional', {}).get('experiencia_relacionada', {}).get('puntaje', ''),
            a.get('evaluacion_motivacional', {}).get('experiencia_relacionada', {}).get('justificacion', ''),
            a.get('evaluacion_motivacional', {}).get('experiencia_relacionada', {}).get('tipo_motivacion', ''),
            a.get('evaluacion_motivacional', {}).get('proyeccion_vida', {}).get('puntaje', ''),
            a.get('evaluacion_motivacional', {}).get('proyeccion_vida', {}).get('justificacion', ''),
            a.get('evaluacion_motivacional', {}).get('proyeccion_vida', {}).get('tipo_motivacion', ''),
            a.get('nivel_motivacional_general', ''),
            a.get('necesidades_psicologicas', {}).get('autonomia', ''),
            a.get('necesidades_psicologicas', {}).get('competencia', ''),
            a.get('necesidades_psicologicas', {}).get('relacion', ''),
            a.get('recomendaciones', '')
        ])
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def main():
    st.markdown("""
    <div class="app-header">
        <h1 class="app-title">Sistema de Analisis de Admision</h1>
        <p class="app-subtitle">Analisis Motivacional basado en la Teoria de la Autodeterminacion (SDT)</p>
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("""
        <div style='background: #FFFFFF; padding: 1.5rem; border-radius: 12px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); text-align: center;'>
            <p style='margin: 0; color: #334155; font-size: 0.95rem;'>
                <strong style='color: #1E3A8A;'>Analisis Individual:</strong> PDF, DOCX, TXT |
                <strong style='color: #1E3A8A;'>Analisis Masivo:</strong> XLSX, XLS, CSV
            </p>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("<br>", unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader("Seleccionar Archivo", type=['pdf', 'docx', 'doc', 'txt', 'xlsx', 'xls', 'csv'], help="Arrastra el archivo o haz clic para seleccionar", label_visibility="collapsed")
    
    if uploaded_file:
        file_extension = uploaded_file.name.split('.')[-1].lower()
        file_size = uploaded_file.size / (1024 * 1024)
        is_batch = file_extension in ['xlsx', 'xls', 'csv']
        
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            st.markdown(f"<div class='info-box'><strong>{uploaded_file.name}</strong></div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div class='info-box'>{file_size:.2f} MB</div>", unsafe_allow_html=True)
        with col3:
            mode = "Modo Masivo" if is_batch else "Modo Individual"
            badge = "badge-info" if is_batch else "badge-success"
            st.markdown(f"<div class='info-box'><span class='status-badge {badge}'>{mode}</span></div>", unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            analyze_button = st.button("Iniciar Analisis", type="primary", use_container_width=True)
        
        if analyze_button:
            with st.spinner("Procesando analisis..."):
                if is_batch:
                    df = read_excel_file(uploaded_file)
                    if df is not None:
                        st.success(f"{len(df)} registros detectados")
                        col_map = build_column_map(df)
                        with st.expander("Mapeo de columnas detectado", expanded=False):
                            for logical, real in col_map.items():
                                st.markdown(f"**{logical}** -> {'`' + real + '`' if real else 'No encontrada'}")
                        
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        results = process_excel_records(df, progress_bar, status_text)
                        status_text.markdown("**Analisis completado**")
                        progress_bar.progress(1.0)
                        
                        st.session_state['batch_results'] = results
                        st.markdown("<hr>", unsafe_allow_html=True)
                        st.markdown("## Resultados del Analisis Masivo")
                        
                        success_count = sum(1 for r in results if r.get('success'))
                        avg_score = sum(float(r['analysis']['calificacion_sobre_20']) for r in results if r.get('success') and r.get('analysis', {}).get('calificacion_sobre_20')) / success_count if success_count > 0 else 0
                        
                        c1, c2, c3, c4 = st.columns(4)
                        c1.metric("Total", len(results))
                        c2.metric("Exitosos", success_count)
                        c3.metric("Errores", len(results) - success_count)
                        c4.metric("Promedio", f"{avg_score:.2f}/20")
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        excel_buffer = generate_excel_report(results)
                        col1, col2, col3 = st.columns([1, 2, 1])
                        with col2:
                            st.download_button("Descargar Reporte Excel", data=excel_buffer, file_name=f"Analisis_SDT_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                        
                        st.markdown("<br>", unsafe_allow_html=True)
                        st.markdown("### Detalle por Postulante")
                        
                        for i, result in enumerate(results):
                            nombre_d = result.get('nombre', 'N/A')
                            apellidos_d = result.get('apellidos', 'N/A')
                            correo_d = result.get('correo', 'N/A')
                            parts = []
                            if apellidos_d != 'N/A': parts.append(apellidos_d)
                            if nombre_d != 'N/A': parts.append(nombre_d)
                            name_label = ", ".join(parts) if parts else f"Registro {result['registro_numero']}"
                            extra = f" | {correo_d}" if correo_d != 'N/A' else ""
                            
                            with st.expander(f"**{result['registro_numero']}. {name_label}**{extra}", expanded=False):
                                if result.get('success'):
                                    analysis = result['analysis']
                                    c1, c2 = st.columns([1, 3])
                                    with c1:
                                        st.markdown(f"<div style='text-align:center;padding:1rem;'><div class='score-circle'>{analysis.get('calificacion_sobre_20', 'N/A')}</div></div>", unsafe_allow_html=True)
                                    with c2:
                                        st.markdown(f"**Calificacion Real:** {analysis.get('calificacion_real', 'N/A')}/18")
                                        st.markdown(f"**Nivel Motivacional:** {analysis.get('nivel_motivacional_general', 'N/A')}")
                                        st.markdown(f"**Regla aplicada:** {analysis.get('regla_aplicada', 'N/A')}")
                                    
                                    st.markdown("---")
                                    st.markdown("#### Evaluacion Motivacional")
                                    eval_mot = analysis.get('evaluacion_motivacional', {})
                                    c1, c2, c3 = st.columns(3)
                                    with c1:
                                        if 'eleccion_carrera' in eval_mot:
                                            st.info(f"**R1: Eleccion**\n\n{eval_mot['eleccion_carrera'].get('puntaje')}/6 - {eval_mot['eleccion_carrera'].get('tipo_motivacion')}")
                                    with c2:
                                        if 'experiencia_relacionada' in eval_mot:
                                            st.info(f"**R2: Experiencia**\n\n{eval_mot['experiencia_relacionada'].get('puntaje')}/6 - {eval_mot['experiencia_relacionada'].get('tipo_motivacion')}")
                                    with c3:
                                        if 'proyeccion_vida' in eval_mot:
                                            st.info(f"**R3: Proyeccion de Vida**\n\n{eval_mot['proyeccion_vida'].get('puntaje')}/6 - {eval_mot['proyeccion_vida'].get('tipo_motivacion')}")
                                    
                                    st.markdown("#### Justificaciones y Reglas Aplicadas")
                                    for key, label in [('eleccion_carrera', 'R1 - Proceso de Eleccion'), ('experiencia_relacionada', 'R2 - Experiencia Personal'), ('proyeccion_vida', 'R3 - Proyeccion de Vida')]:
                                        if key in eval_mot:
                                            st.markdown(f"**{label}:** {eval_mot[key].get('justificacion', 'N/A')}")
                                    
                                    if 'necesidades_psicologicas' in analysis:
                                        st.markdown("#### Necesidades Psicologicas")
                                        nec = analysis['necesidades_psicologicas']
                                        c1, c2, c3 = st.columns(3)
                                        c1.success(f"**Autonomia**\n\n{nec.get('autonomia', 'N/A')}")
                                        c2.success(f"**Competencia**\n\n{nec.get('competencia', 'N/A')}")
                                        c3.success(f"**Relacion**\n\n{nec.get('relacion', 'N/A')}")
                                    
                                    if 'recomendaciones' in analysis:
                                        st.markdown("#### Recomendaciones")
                                        st.info(analysis['recomendaciones'])
                                else:
                                    st.error(f"**Error:** {result.get('error', 'Desconocido')}")
                
                else:
                    text_content = None
                    if file_extension == 'pdf': text_content = extract_text_from_pdf(uploaded_file)
                    elif file_extension in ['docx', 'doc']: text_content = extract_text_from_docx(uploaded_file)
                    elif file_extension == 'txt': text_content = extract_text_from_txt(uploaded_file)
                    
                    if text_content and text_content.strip():
                        st.success(f"Texto extraido ({len(text_content)} caracteres)")
                        analysis = analyze_admission_form(text_content)
                        
                        if analysis.get('success'):
                            st.markdown("<hr>", unsafe_allow_html=True)
                            st.markdown("## Resultado del Analisis Individual")
                            info = analysis.get('informacion_extraida', {})
                            c1, c2, c3 = st.columns(3)
                            c1.info(f"**Nombre**\n\n{info.get('nombre', 'N/A')}")
                            c2.info(f"**Edad**\n\n{info.get('edad', 'N/A')}")
                            c3.info(f"**Programa**\n\n{info.get('programa', 'N/A')}")
                            
                            st.markdown("<br>", unsafe_allow_html=True)
                            c1, c2 = st.columns([1, 2])
                            with c1:
                                st.markdown(f"<div style='text-align:center;padding:2rem;'><div class='score-circle'>{analysis.get('calificacion_sobre_20', 'N/A')}</div><p style='color:#64748B;margin-top:1rem;font-weight:600;'>Calificacion Final</p></div>", unsafe_allow_html=True)
                            with c2:
                                st.metric("Calificacion Real", f"{analysis.get('calificacion_real', 'N/A')}/18")
                                st.metric("Nivel Motivacional", analysis.get('nivel_motivacional_general', 'N/A'))
                            
                            st.markdown("<br>", unsafe_allow_html=True)
                            st.markdown("### Evaluacion Motivacional Detallada")
                            eval_mot = analysis.get('evaluacion_motivacional', {})
                            for key, label in [('eleccion_carrera', 'Proceso de Eleccion'), ('experiencia_relacionada', 'Experiencia Personal'), ('proyeccion_vida', 'Proyeccion de Vida')]:
                                if key in eval_mot:
                                    item = eval_mot[key]
                                    with st.expander(f"**{label}** - Puntaje: {item.get('puntaje')}/6 - {item.get('tipo_motivacion')}", expanded=True):
                                        st.markdown(f"**Justificacion:** {item.get('justificacion')}")
                            
                            if 'necesidades_psicologicas' in analysis:
                                st.markdown("### Necesidades Psicologicas (SDT)")
                                nec = analysis['necesidades_psicologicas']
                                c1, c2, c3 = st.columns(3)
                                c1.success(f"**Autonomia**\n\n{nec.get('autonomia', 'N/A')}")
                                c2.success(f"**Competencia**\n\n{nec.get('competencia', 'N/A')}")
                                c3.success(f"**Relacion**\n\n{nec.get('relacion', 'N/A')}")
                            
                            if 'recomendaciones' in analysis:
                                st.markdown("### Recomendaciones")
                                st.info(analysis['recomendaciones'])
                            
                            st.markdown("<hr>", unsafe_allow_html=True)
                            st.caption(f"**Archivo:** {uploaded_file.name} | **Tokens:** {analysis.get('tokens_used', 'N/A')} | **Procesado:** {datetime.fromisoformat(analysis.get('timestamp')).strftime('%d/%m/%Y %H:%M:%S')}")
                        else:
                            st.error(f"**Error:** {analysis.get('error', 'Desconocido')}")
                    else:
                        st.error("No se pudo extraer texto del archivo")
    
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("<div style='text-align:center;padding:2rem;color:#94A3B8;font-size:0.85rem;'><p>Direccion de Gestion de la Informacion - Universidad Continental</p></div>", unsafe_allow_html=True)

if __name__ == "__main__":
    main()
